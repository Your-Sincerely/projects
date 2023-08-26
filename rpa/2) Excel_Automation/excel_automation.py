from openpyxl import load_workbook,Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font


def read_excel(file_path):
    try:
        # Check the file extension
        if file_path.endswith('.xlsx'):
            # Read from .xlsx file using openpyxl
            workbook = load_workbook(filename=file_path,data_only=True)
            sheet = workbook.active
            data = [[cell for cell in row] for row in sheet.iter_rows(min_row=1, max_col=sheet.max_column, max_row=sheet.max_row, values_only=True)]
        else:
            raise ValueError("Invalid file format. Only .xls and .xlsx are supported.")
        
        return data
    except Exception as e:
        print(f"An error occurred while reading the Excel file: {str(e)}")
        return None

# Usage example
# file_path = 'example.xlsx'  # Replace with your file path
# data = read_excel(file_path)
# print(data)

# -------------------------------------------------------------------------

def write_excel(file_path, data,output_path):
    try:
        if file_path.endswith('.xlsx'):
            # Write to .xlsx file using openpyxl
            workbook = load_workbook(filename=file_path)
            sheet = workbook.active
            for row_index, row_data in enumerate(data):
                for col_index, value in enumerate(row_data):
                    sheet.cell(row=row_index + 1, column=col_index + 1).value = value
            workbook.save(output_path)
        else:
            raise ValueError("Invalid file format. Only .xls and .xlsx are supported.")
            return False
        return True
        print("Data written successfully to the Excel file.")
    except Exception as e:
        print(f"An error occurred while writing to the Excel file: {str(e)}")
        return False

# Usage example
# file_path = 'example.xlsx'  # Replace with your file path
# data = [['Name', 'Age', 'Country'],
#         ['John', 25, 'USA'],
#         ['Emma', 30, 'UK'],
#         ['Liam', 28, 'Canada']]
# write_excel(file_path, data)

# --------------------------------------------------------------------------


def format_excel(file_path,output_path):
    try:

        # Check the file extension
        if file_path.endswith('.xlsx'):
            # Format .xlsx file using openpyxl
            workbook = load_workbook(filename=file_path)
            sheet = workbook.active
            for row in sheet.iter_rows():
                for cell in row:
                    cell.font = Font(name='Calibri',
                                    size=11,
                                    bold=True,
                                    italic=True,
                                    vertAlign=None,
                                    underline='single',
                                    strike=False,
                                    color='00FF00')
                    cell.alignment = Alignment(horizontal='center',
                                                vertical='center',
                                                text_rotation=0,
                                                wrap_text=False,
                                                shrink_to_fit=False,
                                                indent=0)
                    cell.fill = PatternFill(fill_type='solid',
                                            start_color='FFFF00',
                                            end_color='FF000000')
            workbook.save(output_path)
            
        else:
            raise ValueError("Invalid file format. Only .xls and .xlsx are supported.")
            return False
        return True

        print("Excel file formatting completed successfully.")
    except Exception as e:
        print(f"An error occurred while formatting the Excel file: {str(e)}")
        return False

# Usage example
# file_path = 'example.xlsx'  # Replace with your file path
# format_excel(file_path)

# ------------------------------------------------------------------------------

def calculate_column_average(file_path, column_index):
    try:
        workbook = load_workbook(filename=file_path)
        sheet = workbook.active
        
        column_values = []
        for row in sheet.iter_rows(values_only=True):
            cell_value = row[column_index]
            if isinstance(cell_value, (int, float)):
                column_values.append(cell_value)
        
        if column_values:
            average = sum(column_values) / len(column_values)
            return average
        else:
            return None
    
    except Exception as e:
        print(f"An error occurred while calculating the column average: {str(e)}")
        return None

# Usage example
# file_path = 'example.xlsx'  # Replace with your file path
# perform_calculation(file_path)

# --------------------------------------------------------------------------


# --------------test case-------------------

